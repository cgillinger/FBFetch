# demographics.py
#
# Ett verktyg för att hämta demografisk data från Facebook-sidor
# Skapar en Excel-fil där varje sida har en egen flik med demografisk information
#
# Användning:
#   python demographics.py [--output FILNAMN] [--pages SIDA1,SIDA2,...] [--detailed]
#

import os
import sys
import time
import json
import argparse
import logging
import requests
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importera konfigurationsvariabler från config.py
try:
    from config import (
        ACCESS_TOKEN, TOKEN_LAST_UPDATED, API_VERSION, 
        MAX_RETRIES, RETRY_DELAY
    )
except ImportError:
    print("Fel: Kan inte hitta config.py. Se till att filen finns i samma mapp.")
    sys.exit(1)

# Definition av demografiska metriker att hämta och deras perioder
# Viktigt: Olika metriker har olika giltiga perioder enligt Facebook API
DEMOGRAPHIC_METRICS = {
    # Fans-metriker (kräver "lifetime" period)
    "page_fans_city": {
        "name": "Fans per stad",
        "periods": ["lifetime"],
        "description": "Aggregerad data om fans per stad"
    },
    "page_fans_country": {
        "name": "Fans per land",
        "periods": ["lifetime"],
        "description": "Aggregerad data om fans per land"
    },
    "page_fans_gender_age": {
        "name": "Fans per kön och ålder",
        "periods": ["lifetime"],
        "description": "Aggregerad data om fans per kön och åldersgrupp"
    },
    "page_fans_locale": {
        "name": "Fans per språk",
        "periods": ["lifetime"],
        "description": "Aggregerad data om fans per språkinställning"
    },
    "page_fans_online_per_day": {
        "name": "Fans online per dag",
        "periods": ["lifetime"],
        "description": "När fans är online per veckodag"
    },
    "page_fans_online": {
        "name": "Fans online per timme",
        "periods": ["lifetime"],
        "description": "När fans är online per timme på dygnet"
    },
    
    # Räckviddsmetriker (fungerar med dag, vecka, 28 dagar)
    "page_impressions_by_city_unique": {
        "name": "Räckvidd per stad",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som nåtts av innehåll, per stad"
    },
    "page_impressions_by_country_unique": {
        "name": "Räckvidd per land",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som nåtts av innehåll, per land"
    },
    "page_impressions_by_age_gender_unique": {
        "name": "Räckvidd per ålder/kön",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som nåtts av innehåll, per ålder och kön"
    },
    
    # Interaktionsmetriker
    "page_engaged_users": {
        "name": "Engagerade användare",
        "periods": ["day", "week", "days_28"],
        "description": "Unika användare som interagerat med sidan"
    },
    
    # Placering/Enhetsmetriker
    "page_impressions_by_browser_unique": {
        "name": "Räckvidd per webbläsare",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som nåtts per webbläsare"
    },
    "page_impressions_by_device_type_unique": {
        "name": "Räckvidd per enhetstyp",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som nåtts per enhetstyp (desktop, mobil, etc.)"
    },
    
    # Tillgänglig endast för större sidor med minst 100 personer i kategorin
    "page_content_activity_by_age_gender_unique": {
        "name": "Aktivitet per ålder/kön",
        "periods": ["day", "week", "days_28"],
        "description": "Unika personer som interagerat med innehåll per ålder och kön"
    }
}

# Konfigurera loggning
def setup_logging():
    """Konfigurera loggning med datumstämplad loggfil"""
    now = datetime.now()
    log_dir = "logs"
    
    # Skapa loggdirektory om den inte finns
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    # Skapa datumstämplad loggfilnamn
    log_filename = os.path.join(log_dir, f"demographics_{now.strftime('%Y-%m-%d_%H-%M-%S')}.log")
    
    # Konfigurera loggning
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename),
            logging.FileHandler("demographics.log", mode="w"),  # Överskriver tidigare loggfil
            logging.StreamHandler()  # Terminal-utskrift
        ]
    )
    
    logger = logging.getLogger(__name__)
    logger.info(f"Startar loggning till fil: {log_filename}")
    
    return logger

# Konfigurera loggning
logger = setup_logging()

# Räknare för API-anrop
api_call_count = 0
start_time = time.time()

def check_token_expiry():
    """Kontrollera om token snart går ut och varna användaren"""
    try:
        last_updated = datetime.strptime(TOKEN_LAST_UPDATED, "%Y-%m-%d")
        TOKEN_VALID_DAYS = 60  # Meta tokens är vanligtvis giltiga i 60 dagar
        days_since = (datetime.now() - last_updated).days
        days_left = TOKEN_VALID_DAYS - days_since
        
        logger.info(f"🔑 Token skapades för {days_since} dagar sedan ({days_left} dagar kvar till utgång).")
        
        if days_left <= 0:
            logger.error(f"❌ KRITISKT: Din token har gått ut! Skapa en ny token omedelbart.")
            sys.exit(1)
        elif days_left <= 7:
            logger.warning(f"⚠️ VARNING: Din token går ut inom {days_left} dagar! Skapa en ny token snart.")
    except Exception as e:
        logger.error(f"⚠️ Kunde inte tolka TOKEN_LAST_UPDATED: {e}")

def api_request(url, params, retries=MAX_RETRIES):
    """Gör API-förfrågan med återförsök och rate limit-hantering"""
    global api_call_count
    
    for attempt in range(retries):
        try:
            api_call_count += 1
            logger.debug(f"API-anrop {api_call_count}: {url} med parametrar {params}")
            response = requests.get(url, params=params, timeout=30)
            
            # Kontrollera X-App-Usage och X-Ad-Account-Usage headers för bättre rate limiting
            app_usage = response.headers.get('X-App-Usage')
            if app_usage:
                try:
                    usage_data = json.loads(app_usage)
                    call_count = usage_data.get('call_count', 0)
                    total_time = usage_data.get('total_time', 0)
                    total_cputime = usage_data.get('total_cputime', 0)
                    
                    # Om vi närmar oss gränser, vänta
                    if call_count > 80 or total_time > 80 or total_cputime > 80:  # 80% av gränsen
                        wait_time = 60  # Vänta 1 minut
                        logger.warning(f"API-användning hög: {app_usage}. Väntar {wait_time} sekunder...")
                        time.sleep(wait_time)
                except (json.JSONDecodeError, KeyError):
                    logger.debug(f"Kunde inte tolka X-App-Usage: {app_usage}")
            
            # Hantera vanliga HTTP-fel
            if response.status_code == 429:  # Too Many Requests
                retry_after = int(response.headers.get('Retry-After', RETRY_DELAY))
                logger.warning(f"Rate limit nått! Väntar {retry_after} sekunder... (försök {attempt+1}/{retries})")
                time.sleep(retry_after)
                continue
                
            elif response.status_code >= 500:  # Server error
                wait_time = RETRY_DELAY * (2 ** attempt)  # Exponentiell backoff
                logger.warning(f"Serverfel: {response.status_code}. Väntar {wait_time} sekunder... (försök {attempt+1}/{retries})")
                time.sleep(wait_time)
                continue
            
            # För alla HTTP-svarkoder, försök tolka JSON-innehållet
            try:
                json_data = response.json()
                
                # Särskild hantering för 400-fel (Bad Request)
                if response.status_code == 400 and "error" in json_data:
                    error_code = json_data["error"].get("code")
                    error_msg = json_data["error"].get("message", "Okänt fel")
                    
                    # Hantera specifika felkoder
                    if error_code == 4:  # App-specifikt rate limit
                        wait_time = 60 * (attempt + 1)  # Vänta längre för varje försök
                        logger.warning(f"App rate limit: {error_msg}. Väntar {wait_time} sekunder...")
                        time.sleep(wait_time)
                        continue
                        
                    elif error_code == 100 and "valid insights metric" in error_msg:
                        # Detta är ett förväntat fel om metriken inte stöds, returnera felmeddelandet
                        logger.debug(f"Metrik inte tillgänglig: {error_msg}")
                        return json_data
                        
                    elif error_code == 190:  # Ogiltig token
                        logger.error(f"Access token ogiltig: {error_msg}")
                        return None
                
                # Om vi kommer hit och har en icke-200 status, logga felet men returnera ändå JSON-data
                if response.status_code != 200:
                    logger.debug(f"HTTP-fel {response.status_code}: {response.text}")
                    
                    if attempt < retries - 1 and error_code != 100:  # Försök inte igen för metriska fel
                        wait_time = RETRY_DELAY * (2 ** attempt)
                        logger.info(f"Väntar {wait_time} sekunder innan nytt försök... (försök {attempt+1}/{retries})")
                        time.sleep(wait_time)
                        continue
                    
                    # Returnera ändå JSON-data så att anropande funktion kan hantera felet
                    return json_data
                
                # Allt gick bra, returnera data
                return json_data
                
            except json.JSONDecodeError:
                logger.error(f"Kunde inte tolka JSON-svar: {response.text[:100]}")
                if attempt < retries - 1:
                    wait_time = RETRY_DELAY * (2 ** attempt)
                    logger.info(f"Väntar {wait_time} sekunder innan nytt försök... (försök {attempt+1}/{retries})")
                    time.sleep(wait_time)
                    continue
                return None
                
        except requests.RequestException as e:
            logger.error(f"Nätverksfel: {e}")
            if attempt < retries - 1:
                wait_time = RETRY_DELAY * (2 ** attempt)
                logger.info(f"Väntar {wait_time} sekunder innan nytt försök... (försök {attempt+1}/{retries})")
                time.sleep(wait_time)
            else:
                return None
    
    return None

def validate_token(token):
    """Validera att token är giltig och hämta användarbehörigheter"""
    logger.info("Validerar token...")
    url = f"https://graph.facebook.com/{API_VERSION}/debug_token"
    params = {"input_token": token, "access_token": token}
    
    data = api_request(url, params)
    
    if not data or "data" not in data:
        logger.error("❌ Kunde inte validera token")
        return False
        
    if not data["data"].get("is_valid"):
        logger.error(f"❌ Token är ogiltig: {data['data'].get('error', {}).get('message', 'Okänd anledning')}")
        return False
        
    # Kontrollera behörigheter
    permissions = data["data"].get("scopes", [])
    logger.info(f"✅ Token validerad. App ID: {data['data'].get('app_id')}")
    
    # Kolla kritiska behörigheter 
    required_permissions = {"read_insights", "pages_read_engagement"}
    missing_permissions = required_permissions - set(permissions)
    if missing_permissions:
        logger.warning(f"⚠️ Token saknar följande behörigheter: {', '.join(missing_permissions)}")
        logger.warning("Detta kan begränsa vilken data som kan hämtas.")
    else:
        logger.info("✅ Token har alla nödvändiga behörigheter")
    
    return True

def get_page_ids_with_access(token):
    """Hämta alla sidor som token har åtkomst till"""
    logger.info("Hämtar tillgängliga sidor...")
    url = f"https://graph.facebook.com/{API_VERSION}/me/accounts"
    params = {"access_token": token, "limit": 100, "fields": "id,name,category,fan_count"}
    
    pages = []
    next_url = url
    
    while next_url:
        data = api_request(url if next_url == url else next_url, {} if next_url != url else params)
        
        if not data or "data" not in data:
            break
            
        pages.extend(data["data"])
        logger.debug(f"Hittade {len(data['data'])} sidor i denna batch")
        
        # Hantera paginering
        next_url = data.get("paging", {}).get("next")
        if next_url and next_url != url:
            logger.debug(f"Hämtar nästa sida från: {next_url}")
        else:
            break
    
    if not pages:
        logger.warning("Inga sidor hittades. Token kanske saknar 'pages_show_list'-behörighet.")
        return []
    
    # Sortera sidor efter antal fans (högst först)
    pages.sort(key=lambda p: p.get("fan_count", 0), reverse=True)
    
    # Skapa en lista av tupler med (id, namn, kategori, antal fans)
    page_info = [(page["id"], 
                 page["name"], 
                 page.get("category", "Okänd kategori"), 
                 page.get("fan_count", 0)) for page in pages]
    
    logger.info(f"✅ Hittade {len(page_info)} sidor att analysera")
    
    # Visa de 5 största sidorna
    if len(page_info) > 0:
        logger.info("Största sidor:")
        for i, (page_id, name, category, fan_count) in enumerate(page_info[:5], 1):
            logger.info(f"  {i}. {name} ({category}) - {fan_count:,} fans")
    
    return page_info

def get_page_access_token(page_id, system_token):
    """Konvertera systemanvändartoken till en Page Access Token för en specifik sida"""
    logger.debug(f"Hämtar Page Access Token för sida {page_id}...")
    url = f"https://graph.facebook.com/{API_VERSION}/{page_id}"
    params = {
        "fields": "access_token",
        "access_token": system_token
    }
    
    data = api_request(url, params)
    
    if not data or "error" in data or "access_token" not in data:
        error_msg = data.get("error", {}).get("message", "Okänt fel") if data and "error" in data else "Kunde inte hämta token"
        logger.warning(f"⚠️ Kunde inte hämta Page Access Token för sida {page_id}: {error_msg}")
        return None
    
    return data["access_token"]

def get_demographic_data(page_id, page_name, system_token, detailed=False):
    """Hämta demografisk data för en sida med rätt perioder för varje metrik"""
    logger.info(f"Hämtar demografisk data för sida: {page_name} (ID: {page_id})...")
    
    # Först hämta en Page Access Token för denna specifika sida
    page_token = get_page_access_token(page_id, system_token)
    
    if not page_token:
        logger.warning(f"⚠️ Kunde inte hämta Page Access Token för sida {page_name}")
        return {
            "page_id": page_id,
            "page_name": page_name,
            "error": "Kunde inte hämta page access token",
            "data": {}
        }
    
    # Skapa resultatstruktur
    result = {
        "page_id": page_id,
        "page_name": page_name,
        "error": None,
        "data": {}
    }
    
    # Hämta varje demografisk metrik med rätt period
    has_any_data = False
    
    for metric_name, metric_info in DEMOGRAPHIC_METRICS.items():
        result["data"][metric_name] = {
            "display_name": metric_info["name"],
            "description": metric_info["description"],
            "values": {},
            "error": None
        }
        
        metric_success = False
        
        # Prova varje period som är giltig för denna metrik
        for period in metric_info["periods"]:
            if metric_success:
                break  # Hoppa över om vi redan har data
            
            url = f"https://graph.facebook.com/{API_VERSION}/{page_id}/insights"
            params = {
                "access_token": page_token,
                "metric": metric_name,
                "period": period
            }
            
            # För räckviddmetriker, lägg till tidsgränser för att få nyare data om det är en icke-lifetime period
            if period != "lifetime":
                # För nyare data, använd senaste 28 dagarna
                end_date = datetime.now()
                start_date = end_date - timedelta(days=28)
                params["since"] = start_date.strftime("%Y-%m-%d")
                params["until"] = end_date.strftime("%Y-%m-%d")
            
            logger.debug(f"Hämtar {metric_name} för {page_name} med period={period}")
            
            try:
                data = api_request(url, params)
                
                # Analysera svaret med bättre felhantering
                if data and "data" in data and data["data"]:
                    for metric_data in data["data"]:
                        if metric_data["name"] == metric_name:
                            values = metric_data.get("values", [])
                            
                            if values and len(values) > 0:
                                value_data = values[0].get("value", {})
                                
                                # Lagra värdet i resultatet om det inte är tomt
                                if value_data and isinstance(value_data, dict) and len(value_data) > 0:
                                    result["data"][metric_name]["values"] = value_data
                                    result["data"][metric_name]["period"] = period
                                    logger.info(f"  ✓ {metric_info['name']} data hämtad för {page_name} med period={period}")
                                    metric_success = True
                                    has_any_data = True
                                    break
                                elif value_data and not isinstance(value_data, dict):
                                    # Hantera icke-dictionary-värden (t.ex. mätvärden som är nummer)
                                    result["data"][metric_name]["values"] = {"total": value_data}
                                    result["data"][metric_name]["period"] = period
                                    logger.info(f"  ✓ {metric_info['name']} data hämtad för {page_name} med period={period}")
                                    metric_success = True
                                    has_any_data = True
                                    break
                
                # Om inget värde hittades än
                if not metric_success and data and "error" in data:
                    error_msg = data["error"].get("message", "Okänt fel")
                    error_code = data["error"].get("code", "")
                    
                    # Om det är ett felmeddelande om ogiltig metrik, lagra det men fortsätt mjukt
                    if error_code == 100 and "valid insights metric" in error_msg:
                        logger.warning(f"  ✗ Metrik {metric_name} är inte tillgänglig för denna sida, period={period}")
                        result["data"][metric_name]["error"] = f"Metrik inte tillgänglig: {error_msg}"
                    elif "insufficient" in error_msg.lower() or "permission" in error_msg.lower():
                        logger.warning(f"  ✗ Otillräckliga behörigheter för {metric_name} med period={period}: {error_msg}")
                        result["data"][metric_name]["error"] = f"Otillräckliga behörigheter: {error_msg}"
                    else:
                        logger.warning(f"  ✗ Fel vid hämtning av {metric_name} för sida {page_name}: {error_msg}")
                        result["data"][metric_name]["error"] = error_msg
                
            except Exception as e:
                logger.warning(f"  ✗ Undantag vid hämtning av {metric_name} för sida {page_name}: {str(e)}")
                result["data"][metric_name]["error"] = str(e)
        
        # Om vi fortfarande inte har data för denna metrik, logga ett sammanfattande meddelande
        if not metric_success:
            logger.warning(f"  ✗ Kunde inte hämta {metric_info['name']} för sida {page_name} med någon period")
            # Endast uppdatera huvudfelet om vi inte redan har ett fel
            if result["error"] is None and result["data"][metric_name]["error"]:
                result["error"] = f"Fel för {metric_info['name']}: {result['data'][metric_name]['error']}"
    
    # Kontrollera om vi har någon data alls
    if not has_any_data:
        if result["error"] is None:
            result["error"] = "Ingen demografisk data tillgänglig för denna sida"
        logger.warning(f"⚠️ Ingen demografisk data hittades för sida {page_name}")
    else:
        # Beräkna sammanfattande statistik
        total_fans_by_country = sum(result["data"].get("page_fans_country", {}).get("values", {}).values())
        total_fans_by_city = sum(result["data"].get("page_fans_city", {}).get("values", {}).values())
        
        logger.info(f"📊 Sammanfattning för {page_name}:")
        logger.info(f"  - Totalt antal fans från länder: {total_fans_by_country:,}")
        logger.info(f"  - Totalt antal fans från städer: {total_fans_by_city:,}")
        
        # Om vi har köns- och åldersfördelning, visa sammanfattning
        gender_age_data = result["data"].get("page_fans_gender_age", {}).get("values", {})
        if gender_age_data:
            # Beräkna totaler per kön
            gender_totals = {"M": 0, "F": 0, "U": 0}
            for key, value in gender_age_data.items():
                gender = key.split(".")[0]  # Format är "M.13-17", "F.18-24", etc.
                if gender in gender_totals:
                    gender_totals[gender] += value
            
            # Visa könfördelning
            total_with_gender = sum(gender_totals.values())
            if total_with_gender > 0:
                logger.info(f"  - Könsfördelning:")
                male_percent = gender_totals["M"] / total_with_gender * 100 if total_with_gender > 0 else 0
                female_percent = gender_totals["F"] / total_with_gender * 100 if total_with_gender > 0 else 0
                unknown_percent = gender_totals["U"] / total_with_gender * 100 if total_with_gender > 0 else 0
                
                logger.info(f"    - Män: {gender_totals['M']:,} ({male_percent:.1f}%)")
                logger.info(f"    - Kvinnor: {gender_totals['F']:,} ({female_percent:.1f}%)")
                logger.info(f"    - Okänt: {gender_totals['U']:,} ({unknown_percent:.1f}%)")
    
    # Vänta lite mellan anrop för att inte överlasta API:et
    time.sleep(1)
    
    return result

def process_pages(pages, output_file, selected_page_ids=None, detailed=False):
    """Bearbeta alla sidor och skapa en Excel-fil med demografisk data"""
    if selected_page_ids:
        # Filtrera endast valda sidor om en lista angetts
        filtered_pages = [(page_id, name, category, fans) for page_id, name, category, fans in pages 
                          if page_id in selected_page_ids]
        logger.info(f"Filtrerar till {len(filtered_pages)} av {len(pages)} sidor baserat på indata.")
        pages = filtered_pages
    
    logger.info(f"Bearbetar {len(pages)} sidor...")
    results = []
    
    for i, (page_id, page_name, category, fans) in enumerate(pages):
        logger.info(f"Bearbetar sida {i+1}/{len(pages)}: {page_name} ({category}, {fans:,} fans)")
        data = get_demographic_data(page_id, page_name, ACCESS_TOKEN, detailed=detailed)
        data["category"] = category
        data["fans"] = fans
        results.append(data)
        
        # Vänta lite mellan anrop för att inte överbelasta API:et
        if i < len(pages) - 1:
            time.sleep(2)
    
    # Skapa Excel-filen
    create_excel_report(results, output_file, detailed=detailed)
    
    return results

def format_excel_sheet(worksheet, title, start_row=0):
    """Ställ in grundläggande formatering för ett Excel-ark"""
    # Ställ in titeln
    title_cell = worksheet.cell(row=start_row+1, column=1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=5)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Ställ in kolumnbredderna
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    
    # Formatera rubriker
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for row in worksheet.iter_rows(min_row=start_row+3, max_row=start_row+3, min_col=1, max_col=5):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    return worksheet

def create_excel_report(results, output_file, detailed=False):
    """Skapa en Excel-rapport med en flik per sida"""
    logger.info(f"Skapar Excel-rapport: {output_file}")
    
    # Skapa en Excel-writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Kontrollera att utdatakatalogen finns
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)
    
    # Skapa även CSV-export för varje datatyp
    csv_dir = os.path.join(os.path.dirname(output_file), "csv_export")
    if not os.path.exists(csv_dir):
        os.makedirs(csv_dir)
    
    # Samla all data för CSV-export
    all_data_rows = []
    
    # Skapa översiktsflik
    overview_data = []
    for result in results:
        # Sammanställ statistik
        page_name = result["page_name"]
        page_id = result["page_id"]
        fans = result.get("fans", 0)
        category = result.get("category", "Okänd")
        
        # Räkna antalet objekt för varje metrik
        city_fans_count = len(result["data"].get("page_fans_city", {}).get("values", {}))
        country_fans_count = len(result["data"].get("page_fans_country", {}).get("values", {}))
        gender_age_fans_count = len(result["data"].get("page_fans_gender_age", {}).get("values", {}))
        locale_fans_count = len(result["data"].get("page_fans_locale", {}).get("values", {}))
        
        # Beräkna totalsumma av fans från olika måttvärden när det är tillgängligt
        total_country_fans = sum(result["data"].get("page_fans_country", {}).get("values", {}).values())
        total_city_fans = sum(result["data"].get("page_fans_city", {}).get("values", {}).values())
        total_gender_age_fans = sum(result["data"].get("page_fans_gender_age", {}).get("values", {}).values())
        
        # Sammanställ könfördelning om tillgänglig
        gender_age_data = result["data"].get("page_fans_gender_age", {}).get("values", {})
        male_fans = sum(v for k, v in gender_age_data.items() if k.startswith("M."))
        female_fans = sum(v for k, v in gender_age_data.items() if k.startswith("F."))
        unknown_gender_fans = sum(v for k, v in gender_age_data.items() if k.startswith("U."))
        
        # Beräkna procent av könsfördelning
        total_gender = male_fans + female_fans + unknown_gender_fans
        male_percent = (male_fans / total_gender * 100) if total_gender > 0 else 0
        female_percent = (female_fans / total_gender * 100) if total_gender > 0 else 0
        unknown_percent = (unknown_gender_fans / total_gender * 100) if total_gender > 0 else 0
        
        # Skapa översiktsraden
        page_data = {
            "Sida": page_name,
            "ID": page_id,
            "Kategori": category,
            "Deklarerade fans": fans,
            "API Fans (länder)": total_country_fans,
            "API Fans (städer)": total_city_fans,
            "API Fans (kön/ålder)": total_gender_age_fans,
            "Män": male_fans,
            "Män %": f"{male_percent:.1f}%",
            "Kvinnor": female_fans,
            "Kvinnor %": f"{female_percent:.1f}%",
            "Okänt kön": unknown_gender_fans,
            "Okänt kön %": f"{unknown_percent:.1f}%",
            "Antal länder": country_fans_count,
            "Antal städer": city_fans_count,
            "Antal åldersgrupper": gender_age_fans_count,
            "Antal språk": locale_fans_count,
            "Status": "OK" if result["error"] is None else "FEL",
            "Felmeddelande": result["error"] or ""
        }
        overview_data.append(page_data)
    
    # Skapa översiktsfliken
    if overview_data:
        overview_df = pd.DataFrame(overview_data)
        overview_df.to_excel(writer, sheet_name="Översikt", index=False)
        
        # Formatera översiktsfliken
        worksheet = writer.sheets["Översikt"]
        for column in overview_df:
            column_width = max(
                overview_df[column].astype(str).map(len).max(), 
                len(column)
            )
            col_idx = overview_df.columns.get_loc(column) + 1
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(column_width + 2, 30)
    
    # Skapa en flik per sida med detaljerad data
    for result in results:
        page_name = result["page_name"]
        page_id = result["page_id"]
        sheet_name = page_name[:31]  # Excel har en begränsning på 31 tecken för fliknamn
        
        # Om vi har ett dupliktflknamn, lägg till en del av ID:t
        suffix = 1
        original_sheet_name = sheet_name
        while sheet_name in writer.sheets:
            sheet_name = f"{original_sheet_name[:27]}_{suffix}"
            suffix += 1
        
        # Skapa ett DataFrame för sidans basinfo
        info_df = pd.DataFrame({
            "Information": [
                "Sidnamn", 
                "Sid-ID", 
                "Kategori", 
                "Antal fans (deklarerat)", 
                "Status", 
                "Felmeddelande", 
                "Datum för rapport"
            ],
            "Värde": [
                result["page_name"], 
                result["page_id"], 
                result.get("category", "Okänd"),
                result.get("fans", 0),
                "OK" if result["error"] is None else "FEL",
                result["error"] or "",
                datetime.now().strftime("%Y-%m-%d %H:%M")
            ]
        })
        
        # Skriv basinfo till Excel-flik
        info_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
        
        row_offset = 9  # Börja efter infon
        
        # Organisera data i grupper
        data_groups = [
            {"title": "Fans per land", "metric": "page_fans_country", "columns": ["Land", "Antal fans"]},
            {"title": "Fans per stad", "metric": "page_fans_city", "columns": ["Stad", "Antal fans"]},
            {"title": "Fans per kön och ålder", "metric": "page_fans_gender_age", "columns": ["Kön och ålder", "Antal fans"]},
            {"title": "Fans per språk", "metric": "page_fans_locale", "columns": ["Språk", "Antal fans"]},
            {"title": "Fans online per veckodag", "metric": "page_fans_online_per_day", "columns": ["Veckodag", "Antal fans"]},
            {"title": "Fans online per timme", "metric": "page_fans_online", "columns": ["Timme", "Antal fans"]},
            {"title": "Räckvidd per land", "metric": "page_impressions_by_country_unique", "columns": ["Land", "Räckvidd"]},
            {"title": "Räckvidd per stad", "metric": "page_impressions_by_city_unique", "columns": ["Stad", "Räckvidd"]},
            {"title": "Räckvidd per kön och ålder", "metric": "page_impressions_by_age_gender_unique", "columns": ["Kön och ålder", "Räckvidd"]},
            {"title": "Räckvidd per webbläsare", "metric": "page_impressions_by_browser_unique", "columns": ["Webbläsare", "Räckvidd"]},
            {"title": "Räckvidd per enhetstyp", "metric": "page_impressions_by_device_type_unique", "columns": ["Enhetstyp", "Räckvidd"]}
        ]
        
        for group in data_groups:
            # Hämta data för denna metrik
            metric_data = result["data"].get(group["metric"], {})
            metric_values = metric_data.get("values", {})
            metric_error = metric_data.get("error", None)
            metric_period = metric_data.get("period", "")
            
            # Skapa titel
            title = f"{group['title']} ({metric_period})" if metric_period else group["title"]
            title_df = pd.DataFrame([[title]])
            title_df.to_excel(writer, sheet_name=sheet_name, startrow=row_offset, startcol=0, header=False, index=False)
            row_offset += 1
            
            if metric_values:
                # Skapa dataframe och sortera efter värde (högst först)
                df = pd.DataFrame(metric_values.items(), columns=group["columns"])
                
                # För gender_age, konvertera till mer läsbara etiketter
                if group["metric"] == "page_fans_gender_age" or group["metric"] == "page_impressions_by_age_gender_unique":
                    df[group["columns"][0]] = df[group["columns"][0]].apply(lambda x: format_gender_age(x))
                
                # Sortera efter värde högst först
                df = df.sort_values(group["columns"][1], ascending=False)
                
                # Skriv data
                df.to_excel(writer, sheet_name=sheet_name, startrow=row_offset, startcol=0, index=False)
                
                # Beräkna totalsumma
                total = sum(metric_values.values())
                total_df = pd.DataFrame([["TOTALT", total]], columns=group["columns"])
                total_df.to_excel(writer, sheet_name=sheet_name, startrow=row_offset+len(df)+1, startcol=0, header=False, index=False)
                
                # Uppdatera rad-offset för nästa grupp
                row_offset += len(df) + 4
                
                # Lägg till data för CSV-export
                for key, value in metric_values.items():
                    formatted_key = format_gender_age(key) if group["metric"] in ["page_fans_gender_age", "page_impressions_by_age_gender_unique"] else key
                    all_data_rows.append({
                        "Page name": page_name,
                        "Page ID": page_id,
                        "Dimension": group["metric"],
                        "Category": group["title"],
                        "Key": formatted_key,
                        "Value": value,
                        "Period": metric_period
                    })
            else:
                # Skapa "ingen data" meddelande
                no_data_message = f"Ingen data tillgänglig: {metric_error}" if metric_error else "Ingen data tillgänglig"
                no_data_df = pd.DataFrame([[no_data_message]])
                no_data_df.to_excel(writer, sheet_name=sheet_name, startrow=row_offset, startcol=0, header=False, index=False)
                
                row_offset += 3
            
            # Lägg till lite extra rader mellan grupperna
            row_offset += 1
        
        logger.debug(f"Skapade flik för {page_name}")
    
    # Spara Excel-filen
    writer.close()
    logger.info(f"✅ Excel-rapport sparad till {output_file}")
    
    # Skapa CSV-export
    if all_data_rows:
        csv_path = os.path.join(csv_dir, "demographic_full_export.csv")
        csv_df = pd.DataFrame(all_data_rows)
        csv_df.to_csv(csv_path, index=False, encoding='utf-8')
        logger.info(f"✅ CSV-export sparad till {csv_path}")
        
        # Skapa specifika exports per dimensionstyp
        for dimension in set(row["Dimension"] for row in all_data_rows):
            dimension_data = [row for row in all_data_rows if row["Dimension"] == dimension]
            if dimension_data:
                clean_dimension = dimension.replace("page_", "").replace("_", "-")
                dimension_csv_path = os.path.join(csv_dir, f"demographic_{clean_dimension}.csv")
                dimension_df = pd.DataFrame(dimension_data)
                dimension_df.to_csv(dimension_csv_path, index=False, encoding='utf-8')
                logger.info(f"✅ CSV för {clean_dimension} sparad till {dimension_csv_path}")

def format_gender_age(key):
    """Formatera kön och åldersnycklar till läsbara etiketter"""
    if not isinstance(key, str) or "." not in key:
        return key
        
    gender, age_range = key.split(".", 1)
    
    gender_map = {
        "M": "Man",
        "F": "Kvinna",
        "U": "Okänt kön"
    }
    
    gender_text = gender_map.get(gender, gender)
    return f"{gender_text}, {age_range}"

def parse_args():
    """Parsa kommandoradsargument"""
    parser = argparse.ArgumentParser(description="Hämta demografisk data för Facebook-sidor")
    parser.add_argument("--output", help="Filnamn för Excel-rapporten (standard: fb_demographics.xlsx)", default="fb_demographics.xlsx")
    parser.add_argument("--pages", help="Kommaseparerad lista med sida-ID:n (om tom, hämtas alla tillgängliga sidor)")
    parser.add_argument("--detailed", action="store_true", help="Hämta detaljerad data (tar längre tid)")
    parser.add_argument("--debug", action="store_true", help="Aktivera utförlig loggning")
    return parser.parse_args()

def main():
    """Huvudfunktion"""
    args = parse_args()
    
    # Sätt debug-läge om begärt
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
        logger.debug("Debug-läge aktiverat")
    
    logger.info(f"📊 Facebook Demographics Reporter – v2.0")
    logger.info(f"Output-fil: {args.output}")
    logger.info("-------------------------------------------------------------------")
    
    # Kontrollera token och varna om den snart går ut
    check_token_expiry()
    
    # Validera token
    if not validate_token(ACCESS_TOKEN):
        logger.error("❌ Token kunde inte valideras. Avbryter.")
        return
    
    # Lista med specifika page IDs om angivna
    selected_page_ids = None
    if args.pages:
        selected_page_ids = [page_id.strip() for page_id in args.pages.split(",")]
        logger.info(f"Kommer endast att hämta data för {len(selected_page_ids)} specifika sidor")
    
    # Hämta sidor att bearbeta
    page_list = get_page_ids_with_access(ACCESS_TOKEN)
    
    if not page_list:
        logger.error("❌ Inga sidor hittades. Avbryter.")
        return
    
    # Bearbeta sidor och skapa rapport
    results = process_pages(page_list, args.output, selected_page_ids=selected_page_ids, detailed=args.detailed)
    
    # Visa statistik om API-användning
    elapsed_time = time.time() - start_time
    logger.info(f"⏱️ Total körtid: {elapsed_time:.1f} sekunder")
    logger.info(f"🌐 API-anrop: {api_call_count} ({api_call_count/elapsed_time*3600:.1f}/timme)")
    logger.info(f"✅ Klar! Bearbetade {len(results)} sidor")
    
    # Visa sammanfattade resultat
    successful_pages = sum(1 for r in results if r["error"] is None)
    logger.info(f"📊 Lyckades hämta data för {successful_pages} av {len(results)} sidor")
    
    # Visa några sammanfattande felorsaker om relevanta
    error_types = {}
    for r in results:
        if r["error"]:
            # Extrahera grundorsak från felmeddelandet
            error_type = "Annan fel"
            if "not available" in r["error"].lower() or "valid insights metric" in r["error"].lower():
                error_type = "Metrik ej tillgänglig"
            elif "permission" in r["error"].lower():
                error_type = "Behörighetsfel"
            elif "token" in r["error"].lower():
                error_type = "Token-problem"
            elif "rate limit" in r["error"].lower():
                error_type = "Rate limit"
            
            error_types[error_type] = error_types.get(error_type, 0) + 1
    
    if error_types:
        logger.info("Vanliga felorsaker:")
        for error_type, count in error_types.items():
            logger.info(f"  - {error_type}: {count} sidor")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logger.info("Avbruten av användare.")
        sys.exit(1)
    except Exception as e:
        logger.critical(f"Oväntat fel: {e}")
        import traceback
        logger.critical(traceback.format_exc())
        sys.exit(1)
